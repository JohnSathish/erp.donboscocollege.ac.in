from pathlib import Path
import csv
from openpyxl import Workbook
from fpdf import FPDF

entries = [
    ("Metadata", "Unique ID", "7JBVZT"),
    ("Metadata", "Admission No", "DBCT0039"),
    ("Metadata", "Programme", "FOUR YEAR UNDERGRADUATE PROGRAMME (FYUP) SEMESTER I"),
    ("Personal Information", "Name of the Candidate", "GYAMAR YARA"),
    ("Personal Information", "Date of Birth", "16.10.2009"),
    ("Personal Information", "Gender", "Female"),
    ("Personal Information", "Shift", "SHIFT - II (TIMING : 9.45 AM - 3.30 PM)"),
    ("Personal Information", "Marital Status", "Single"),
    ("Personal Information", "Blood Group", "O+"),
    ("Personal Information", "Category", "ST"),
    ("Personal Information", "Race / Tribe", "Nyishi"),
    ("Personal Information", "Religion", "Christian"),
    ("Personal Information", "Aadhar Number", "412904172955"),
    ("Personal Information", "Email ID", "YARAGYAMAR9@GMAIL.COM"),
    ("Personal Information", "Mobile Number", "7085958818"),
    ("Personal Information", "Is Differently Abled?", "No"),
    ("Personal Information", "Economically Weaker", "No"),
    ("Personal Information", "Brother/Sister Studying in this College (Roll No.)", "No"),
    ("Addresses", "Address in Tura", "TURA QUATER NO. 87 C SECTOR, ARUNACHAL PRADESH"),
    ("Addresses", "Home Address", "NAHARLAGUN, ARUNACHAL PRADESH"),
    ("Parents", "Father Name", "GYAMAR TAMA"),
    ("Parents", "Father Age", "56"),
    ("Parents", "Father Education", "N/A"),
    ("Parents", "Father Occupation", "MTS"),
    ("Parents", "Father Contact No. 1", "9402934932"),
    ("Parents", "Father Contact No. 2", "N/A"),
    ("Parents", "Mother Name", "GYAMAR MOI"),
    ("Parents", "Mother Age", "50"),
    ("Parents", "Mother Education", "N/A"),
    ("Parents", "Mother Occupation", "MTS"),
    ("Parents", "Mother Contact No. 1", "8787789179"),
    ("Parents", "Mother Contact No. 2", "N/A"),
    ("Parents", "Local Guardian Name", "N/A"),
    ("Residence", "Whether Belonging to Urban / Rural?", "Urban"),
    ("Class XII Subjects", "English", "48"),
    ("Class XII Subjects", "History", "42"),
    ("Class XII Subjects", "Political Science", "45"),
    ("Class XII Subjects", "Information Technology", "47"),
    ("Class XII Subjects", "Economics", "46"),
    ("Class XII Subjects", "Geography", "34"),
    ("Class XII Subjects", "Total", "262"),
    ("Exam Details", "Roll No", "16609384"),
    ("Exam Details", "Year", "2025"),
    ("Exam Details", "% of Marks", "43.6"),
    ("Exam Details", "Division", "Third"),
    ("Exam Details", "Name of Board/University", "Central Board of Secondary Education"),
    ("Exam Details", "Regular / Private", "Regular"),
    ("Exam Details", "Marks Scored in CUET", "134"),
    ("Exam Details", "CUET Roll No", "AL0"),
    ("Previous Institution", "Name of Institution Last Attended", "Government Higher Secondary School, Kankarnallah, Naharlagun"),
    ("Previous Institution", "Period of Study From", "N/A"),
    ("Previous Institution", "Period of Study To", "N/A"),
    ("Previous Institution", "Place", "N/A"),
    ("Course Selection", "Major Subject", "English"),
    ("Course Selection", "Subjects as Minor", "Education"),
    ("Course Selection", "MDC (First Semester)", "MDC 112 - Fundamentals of Computer Systems"),
    ("Course Selection", "AEC (First Semester)", "AEC 120 - Alternative English"),
    ("Course Selection", "SEC (First Semester)", "SEC 132 - Personality Development"),
    ("Course Selection", "VAC (First Semester)", "VAC 140 - Environment Studies (Compulsory)"),
    ("Additional", "Admission Form Number", "DBCT0039"),
]

docs_path = Path("docs")
docs_path.mkdir(exist_ok=True)

# CSV export
csv_path = docs_path / "DBCT0039_data.csv"
with csv_path.open("w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Section", "Field", "Value"])
    writer.writerows(entries)

# Excel export
wb = Workbook()
ws = wb.active
ws.title = "DBCT0039"
ws.append(["Section", "Field", "Value"])
for row in entries:
    ws.append(list(row))
excel_path = docs_path / "DBCT0039_data.xlsx"
wb.save(excel_path)

# PDF export
pdf_path = docs_path / "DBCT0039_data.pdf"
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
pdf.set_font("Arial", "B", 14)
pdf.cell(0, 10, "DBCT0039 Data Summary", ln=True)
pdf.ln(4)

current_section = None
pdf.set_font("Arial", "", 11)
for section, field, value in entries:
    if section != current_section:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, section, ln=True)
        pdf.set_font("Arial", "", 11)
        current_section = section
    pdf.multi_cell(0, 6, f"{field}: {value}")
    pdf.ln(1)

pdf.output(str(pdf_path))

print("Exported:")
print(f" - {csv_path}")
print(f" - {excel_path}")
print(f" - {pdf_path}")
